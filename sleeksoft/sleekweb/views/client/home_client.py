from ...models import *

from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_list_or_404, get_object_or_404
from django.core.paginator import Paginator


from django.http import HttpResponse
import requests
import time

from django.db import models
from django.utils import timezone

import os

from datetime import datetime

from django.shortcuts import redirect
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login, logout

from django.contrib.postgres.search import TrigramSimilarity
from django.db.models import Q
from django.shortcuts import render, redirect, reverse
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from datetime import datetime
from django.contrib import messages
import random
import string
from django.contrib.auth import update_session_auth_hash
from datetime import datetime, timedelta
from django.utils.timezone import make_aware

# from PIL import Image, ImageDraw, ImageFont
import requests
from io import BytesIO

import random
import string

import base64

import time
from django.http import JsonResponse

import re
import json

from django.conf import settings
from django.db.models import Q

import datetime

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


import base64

from django.core.mail import send_mail
from django.forms.models import model_to_dict
from django.core.mail import send_mail,EmailMessage

import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import time
import random


def set_language(request, lang_code):
    response = redirect(request.META.get('HTTP_REFERER', '/'))  # quay lại trang vừa bấm
    response.set_cookie('language', lang_code, max_age=60*60*24*30)  # 30 ngày
    return response

def check_count_keyword(request,count_keyword):
    try:
        obj = Price_list.objects.get(Order=1)
        Keyword_day = obj.Keyword_day
        Keyword_month = Keyword_day*30
    except:
        pass
    # Lấy thời gian hiện tại
    now = timezone.now()
    today = now.date()

    try:
        obj_Keyword = request.user.obj_Keyword
    except:
        obj_Keyword = Keyword.objects.create(Belong_User=request.user)

    # So sánh ngày/tháng/năm
    if obj_Keyword.Now_time.date() == today:
        if obj_Keyword.Count_keyword >= Keyword_day:
            messages.error(request, 'Hết giới hạn tìm kiếm trong ngày.')
            return redirect('home_client')
        else:
            obj_Keyword.Count_keyword += count_keyword
    else:
        # Nếu khác ngày, reset Count và cập nhật ngày mới
        obj_Keyword.Now_time = now
        obj_Keyword.Count_keyword = count_keyword

    obj_Keyword.save()

    # Tiếp tục xử lý
    return True

from googlesearch import search

def check_rank(domain, keyword, max_results):
    proxy = ''
    for index, url in enumerate(search(keyword, num_results=max_results)):
        if domain in url:
            return index + 1
    return None

from pathlib import Path
import os
# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent
from serpapi import GoogleSearch
import environ

env = environ.Env()
# environ.Env.read_env()
environ.Env.read_env(os.path.join(BASE_DIR, '.env'))


def get_rank_serpapi(keyword,device,domain):
    try:
        obj_key = API_key.objects.get(Order=1)
    except:
        obj_key = API_key.objects.create(Order=1)
    params = {
        "q": keyword,
        "num": 100,
        "api_key": obj_key.Key,
        "engine": "google",
        "hl": "vi",
        "gl" : "vn",
        "device": device
    }

    search = GoogleSearch(params)
    results = search.get_dict()

    if "organic_results" not in results:
        print("Không có kết quả trả về.")
        return 'Không xác định'

    for idx, item in enumerate(results["organic_results"], start=1):
        link = item.get("link", "")
        if domain in link:
            return idx


def check_rank_keyword(non_empty_lines,device,domain):
    data_check_rank_keyword = []
    for i in non_empty_lines:
        result = get_rank_serpapi(i,device,domain)
        obj = {
            "keyword" : i,
            "rank" : result
        }
        data_check_rank_keyword.append(obj)
    return data_check_rank_keyword

def rank_summary_list(data_check_rank_keyword):
    result = {
        "Top 3": 0,
        "Top 10": 0,
        "Top 30": 0,
        "Top 100": 0,
        "Out of Top 100": 0
    }

    for item in data_check_rank_keyword:
        try:
            rank = int(item["rank"])
            print('rank:',rank)
            if rank is None:
                result["Out of Top 100"] += 1
            elif rank <= 3:
                result["Top 3"] += 1
                result["Top 10"] += 1
                result["Top 30"] += 1
                result["Top 100"] += 1
            elif rank <= 10:
                result["Top 10"] += 1
                result["Top 30"] += 1
                result["Top 100"] += 1
            elif rank <= 30:
                result["Top 30"] += 1
                result["Top 100"] += 1
            elif rank <= 100:
                result["Top 100"] += 1
            else:
                result["Out of Top 100"] += 1
        except:
            result["Out of Top 100"] += 1
    # Chuyển thành list gồm 5 dict như yêu cầu
    summary_list = [
        {"name": "Top 3", "value": result["Top 3"]},
        {"name": "Top 10", "value": result["Top 10"]},
        {"name": "Top 30", "value": result["Top 30"]},
        {"name": "Top 100", "value": result["Top 100"]},
        # {"name": "Out of Top 100", "value": result["Out of Top 100"]}
    ]
    return summary_list


def get_proxy():
    return ''

def home_client(request):
    if request.method == 'GET':
        context = {}
        lc = request.COOKIES.get('language') or 'en'
        context['lc'] = lc
        context['domain'] = settings.DOMAIN
        context['message'] = f"Hello, I saw profile on https://{settings.DOMAIN}"
        try:
            context['obj_key'] = API_key.objects.get(Order=1)
        except:
            context['obj_key'] = API_key.objects.create(Order=1)
        print('key:',context['obj_key'].Key)
        # context['list_Product'] = Product.objects.all()
        # print('context:',context)
        # p = request.GET.get('p','1')
        # context['p'] = p
        # # Sử dụng Paginator để chia nhỏ danh sách (10 là số lượng mục trên mỗi trang)
        # paginator = Paginator(context['list_Product'], settings.PAGE)
        # # Lấy số trang hiện tại từ URL, nếu không mặc định là trang 1
        # context['list_Product'] = paginator.get_page(p)
        try:
            obj = Price_list.objects.get(Order=1)
            context['Price_one'] = obj.Price_one
            context['Price_month'] = obj.Price_month
            context['Keyword_day'] = obj.Keyword_day
            context['Keyword_month'] = context['Keyword_day']*30
        except:
            pass
        context['data_check_rank_keyword'] = request.session.get('data_check_rank_keyword', [])
        context['Device'] = request.session.get('Device', None)
        context['Domain'] = request.session.get('Domain', None)
        context['rank_summary_list'] = request.session.pop('rank_summary_list',
                                                           [
                {"name": "Top 3", "value": 0},
                {"name": "Top 10", "value": 0},
                {"name": "Top 30", "value": 0},
                {"name": "Top 100", "value": 0},
                # {"name": "Out of Top 100", "value": result["Out of Top 100"]}
            ]
                                                            )
        return render(request, 'sleekweb/client/home_client.html', context, status=200)
    if request.method == 'POST':
        if not request.user.is_authenticated:
            messages.error(request, 'Đăng nhập tài khoản trước khi tìm kiếm.')
            return redirect('login_client')
        
        Device = request.POST.get('Device')
        Domain = request.POST.get('Domain')
        Content_keyword = request.POST.get('Content_keyword')

        print('Domain:',Domain)
        print('Content_keyword:',Content_keyword)

        # Tách các dòng, xoá khoảng trắng đầu/cuối, và loại bỏ dòng rỗng
        lines = Content_keyword.strip().splitlines()
        non_empty_lines = [line for line in lines if line.strip() != '']
        print('non_empty_lines:',non_empty_lines)
        # Đếm số dòng có nội dung
        count_keyword = len(non_empty_lines)

        Condition_check_count_keyword = check_count_keyword(request,count_keyword)
        try:
            Condition_time_user = request.user.obj_Time_user.days_left()
        except:
            obj_Time_user = Time_user.objects.create(Belong_User=request.user)
            Condition_time_user = obj_Time_user.days_left()

        print('Condition_time_user:',Condition_time_user)
        proxy = get_proxy()
        if Condition_check_count_keyword:
            if int(Condition_time_user) > 0:
                data_check_rank_keyword = check_rank_keyword(non_empty_lines,Device,Domain)
                Transaction_history.objects.create(
                    Code='SerGoogle',
                    Content='Tiềm kiếm thứ hạng',
                    Belong_User=request.user,
                    Status=1,
                    Value = f"-0"
                    )
            else:
                try:
                    obj_Price_list = Price_list.objects.get(Order=1)
                except:
                    obj_Price_list={}
                print('request.user.Money:',request.user.Money)
                print('obj_Price_list.Price_one:',obj_Price_list.Price_one)
                if request.user.Money <  obj_Price_list.Price_one:
                    print('request.user.Money:','hjghgh')
                    messages.error(request, 'Số tiền không đủ để thực hiện tìm kiếm.')
                    return redirect('home_client')
                data_check_rank_keyword = check_rank_keyword(non_empty_lines,Device,Domain)
                Transaction_history.objects.create(
                    Code='SerGoogle',
                    Content='Tiềm kiếm thứ hạng',
                    Belong_User=request.user,
                    Status=1,
                    Value = f"-{obj_Price_list.Price_one}"
                    )
                request.user.Money = request.user.Money - obj_Price_list.Price_one
                request.user.save()
        request.session['data_check_rank_keyword'] = data_check_rank_keyword
        request.session['Device'] = Device
        request.session['Domain'] = Domain
        request.session['rank_summary_list'] = rank_summary_list(data_check_rank_keyword)
        return redirect('home_client')
    
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Alignment

def export_excel(request):
    if request.method == 'GET':
        data_check_rank_keyword = request.session.get('data_check_rank_keyword', [])
        Device = request.session.get('Device', None)
        Domain = request.session.get('Domain', None)

        now = datetime.now()
        thoi_gian_text = now.strftime("%d-%m-%Y %H:%M:%S")
        ten_file = now.strftime("sergoogle_%d_%m_%Y.xlsx")

        wb = Workbook()
        ws = wb.active

        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 70

        # Dòng 1: Tiêu đề chính (merge + căn giữa)
        ws.merge_cells("A1:B1")
        ws["A1"] = f"Danh sách xếp hạng từ khoá thuộc tên miền '{Domain}', thiết bị '{Device}', thời gian ({thoi_gian_text})"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Dòng 2: Tiêu đề cột
        ws["A2"] = "Từ khóa"
        ws["B2"] = "Xếp hạng"
        ws["A2"].font = Font(bold=True)
        ws["B2"].font = Font(bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["B2"].alignment = Alignment(horizontal="center")

        # Dữ liệu từ dòng 3, căn giữa hết
        for idx, item in enumerate(data_check_rank_keyword, start=3):
            ws[f"A{idx}"] = item['keyword']
            ws[f"B{idx}"] = item['rank']
            ws[f"A{idx}"].alignment = Alignment(horizontal="center")
            ws[f"B{idx}"].alignment = Alignment(horizontal="center")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{ten_file}"'
        return response


